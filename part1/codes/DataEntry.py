import json

import mysql.connector

import pandas as pd

import openpyxl

digikala = mysql.connector.connect(host='127.0.0.1', user='root', password='0000')

cursor = digikala.cursor()

# Creating a database on MySQL Server
cursor.execute('CREATE DATABASE DigiKala')

digikala = mysql.connector.connect(host='127.0.0.1', user='root', password='0000', database='DigiKala', charset='utf8')

cursor = digikala.cursor()

# Creating tables
cursor.execute('CREATE TABLE Book(id INT PRIMARY KEY, name_fa VARCHAR(160), name_en VARCHAR(160), brand_name_fa VARCHAR(64), url_code VARCHAR(160), number_of_copies SMALLINT DEFAULT NULL, category VARCHAR(16) DEFAULT NULL, isbn VARCHAR(18) DEFAULT NULL, author VARCHAR(40) DEFAULT NULL, translator VARCHAR(32) DEFAULT NULL, age_rating VARCHAR(20) DEFAULT NULL, number_of_pages SMALLINT DEFAULT NULL, weight SMALLINT DEFAULT NULL, publisher VARCHAR(16) DEFAULT NULL, cover_size VARCHAR(8) DEFAULT NULL, cover_material VARCHAR(16) DEFAULT NULL, description VARCHAR(80) DEFAULT NULL)')
cursor.execute('CREATE TABLE Puzzle(id INT PRIMARY KEY, name_fa VARCHAR(160), name_en VARCHAR(160), brand_name_fa VARCHAR(64), url_code VARCHAR(160), number_of_pieces VARCHAR(8) DEFAULT NULL, type VARCHAR(8) DEFAULT NULL, age_rating VARCHAR(8) DEFAULT NULL, danger_of_swallow VARCHAR(8) DEFAULT NULL, weight SMALLINT DEFAULT NULL, package_weight SMALLINT DEFAULT NULL, dimensions VARCHAR(20) DEFAULT NULL, package_dimensions VARCHAR(20) DEFAULT NULL, manufacturer VARCHAR(16) DEFAULT NULL, items_in_package VARCHAR(32) DEFAULT  NULL, description VARCHAR(64) DEFAULT NULL)')
cursor.execute('CREATE TABLE Mouse(id INT PRIMARY KEY, name_fa VARCHAR(160), name_en VARCHAR(160), brand_name_fa VARCHAR(64), url_code VARCHAR(160), dimensions VARCHAR(20) DEFAULT NULL, weight SMALLINT DEFAULT NULL, number_of_buttons VARCHAR(8) DEFAULT NULL, power_button VARCHAR(8) DEFAULT NULL, color VARCHAR(8) DEFAULT NULL, two_handed VARCHAR(3) DEFAULT NULL, connection_type VARCHAR(8) DEFAULT NULL, connection_port VARCHAR(8) DEFAULT NULL, cable_material VARCHAR(8) DEFAULT NULL, cable_length VARCHAR(8) DEFAULT NULL, sensor_type VARCHAR(8) DEFAULT NULL, accuracy SMALLINT DEFAULT NULL, accuracy_interval VARCHAR(16) DEFAULT NULL, compatible_os VARCHAR(20) DEFAULT NULL, other_capabilities VARCHAR(32) DEFAULT NULL)')
cursor.execute('CREATE TABLE Keyboard(id INT PRIMARY KEY, name_fa VARCHAR(160), name_en VARCHAR(160), brand_name_fa VARCHAR(64), url_code VARCHAR(160), dimensions VARCHAR(24) DEFAULT NULL, weight SMALLINT DEFAULT NULL, number_of_buttons VARCHAR(8) DEFAULT NULL, keys_lifetime VARCHAR(16) DEFAULT NULL, power_button VARCHAR(8) DEFAULT NULL, dust_proof VARCHAR(3) DEFAULT NULL, persian_letters VARCHAR(5) DEFAULT NULL, backlight VARCHAR(5) DEFAULT NULL, has_mouse VARCHAR(3) DEFAULT NULL, has_touchpad VARCHAR(5) DEFAULT NULL, two_handed VARCHAR(3) DEFAULT NULL, connection_type VARCHAR(8) DEFAULT NULL, connection_port VARCHAR(8) DEFAULT NULL, power_source VARCHAR(3) DEFAULT NULL, cable_length VARCHAR(8) DEFAULT NULL, accuracy VARCHAR(8) DEFAULT NULL, accuracy_interval VARCHAR(16) DEFAULT NULL)')
cursor.execute('CREATE TABLE ScreenProtector(id INT PRIMARY KEY, name_fa VARCHAR(160), name_en VARCHAR(160), brand_name_fa VARCHAR(64), url_code VARCHAR(160), device_model VARCHAR(40) DEFAULT NULL, type VARCHAR(8) DEFAULT NULL, thickness VARCHAR(3) DEFAULT NULL, impact_protection VARCHAR(5) DEFAULT NULL, scratch_protection VARCHAR(5) DEFAULT NULL, anti_reflection VARCHAR(5) DEFAULT NULL, easy_install VARCHAR(5) DEFAULT NULL, protection_part VARCHAR(16) DEFAULT NULL, description VARCHAR(96) DEFAULT NULL)')
cursor.execute('CREATE TABLE Cover(id INT PRIMARY KEY, name_fa VARCHAR(160), name_en VARCHAR(160), brand_name_fa VARCHAR(64), url_code VARCHAR(160), device_model VARCHAR(40) DEFAULT NULL, type VARCHAR(12) DEFAULT NULL, structure VARCHAR(12) DEFAULT NULL, material VARCHAR(12) DEFAULT NULL, weight SMALLINT DEFAULT NULL, dimensions VARCHAR(12) DEFAULT NULL, top_edge_protection VARCHAR(3) DEFAULT NULL, bottom_edge_protection VARCHAR(3) DEFAULT NULL, right_edge_protection VARCHAR(3) DEFAULT NULL, left_edge_protection VARCHAR(3) DEFAULT NULL, back_panel_protection VARCHAR(3) DEFAULT NULL, buttons_protection VARCHAR(3) DEFAULT NULL, special_capabilities VARCHAR(48) DEFAULT NULL, description VARCHAR(320) DEFAULT NULL)')
cursor.execute('CREATE TABLE Brands_FA_EN(brand_name_fa VARCHAR(64) PRIMARY KEY, brand_name_en VARCHAR(64) NOT NULL)')
cursor.execute('CREATE TABLE Buying_History(id INT AUTO_INCREMENT PRIMARY KEY, order_id INT NOT NULL, customer_id INT NOT NULL, item_id INT NOT NULL, cart_finalize_datetime DATETIME, order_gross_amount INT NOT NULL, city VARCHAR(20) NOT NULL, item_quantity INT NOT NULL)')


# Specifying file paths
product_list_path = 'C:\\Users\\Ali\\Desktop\\project\\part1\\codes\\data\\5-awte8wbd.xlsx'
buying_history_path = 'C:\\Users\\Ali\\Desktop\\project\\part1\\codes\\data\\3-p5s3708k.csv'

# Loading Excel file
workbook = openpyxl.load_workbook(product_list_path)
sheet = workbook.active

brands_fa = []
brands_en = []

# Storing Excel file into the database
for i in range(2, 100001):
    category = sheet.cell(row=i, column=6).value
    product_id = sheet.cell(row=i, column=1).value
    name_fa = sheet.cell(row=i, column=2).value
    name_en = sheet.cell(row=i, column=3).value
    url_code = sheet.cell(row=i, column=4).value
    brand_name_fa = sheet.cell(row=i, column=8).value
    brand_name_en = sheet.cell(row=i, column=9).value
    attributes = sheet.cell(row=i, column=10).value
    if attributes:
        attributes = json.loads(attributes)
    if name_en and name_en not in ('NULL', '/', '.'):
        product_values = (product_id, name_fa, name_en, brand_name_fa, url_code)
    else:
        product_values = (product_id, name_fa, None, brand_name_fa, url_code)

    if brand_name_fa not in brands_fa and brand_name_en not in brands_en:
        brands_fa.append(brand_name_fa)
        brands_en.append(brand_name_en)
        cursor.execute('INSERT INTO Brands_FA_EN VALUES (%s, %s)', (brand_name_fa, brand_name_en))

    if category == 'کتاب چاپی':
        cursor.execute('INSERT INTO Book (id, name_fa, name_en, brand_name_fa, url_code) VALUES (%s, %s, %s, %s, %s)', product_values)
        for attribute in attributes:
            if 'Value' in attribute and attribute['Value']:
                if attribute['Key'] == 'تعداد جلد':
                    cursor.execute('UPDATE Book SET number_of_copies = %s WHERE id = %s', (attribute['Value'], product_id))
                elif attribute['Key'] == 'موضوع':
                    cursor.execute('UPDATE Book SET category = %s WHERE id = %s', (attribute['Value'], product_id))
                elif attribute['Key'] == 'شابک':
                    cursor.execute('UPDATE Book SET isbn = %s WHERE id = %s', (attribute['Value'], product_id))
                elif attribute['Key'] == 'نویسنده/نویسندگان':
                    cursor.execute('UPDATE Book SET author = %s WHERE id = %s', (attribute['Value'], product_id))
                elif attribute['Key'] == 'مترجم' and attribute['Value'] != '-':
                    cursor.execute('UPDATE Book SET translator = %s WHERE id = %s', (attribute['Value'], product_id))
                elif attribute['Key'] == 'مناسب برای':
                    cursor.execute('UPDATE Book SET age_rating = %s WHERE id = %s', (attribute['Value'], product_id))
                elif attribute['Key'] == 'تعداد صفحات':
                    cursor.execute('UPDATE Book SET number_of_pages = %s WHERE id = %s', (attribute['Value'], product_id))
                elif attribute['Key'] == 'وزن':
                    cursor.execute('UPDATE Book SET weight = %s WHERE id = %s', (attribute['Value'], product_id))
                elif attribute['Key'] == 'ناشر':
                    cursor.execute('UPDATE Book SET publisher = %s WHERE id = %s', (attribute['Value'], product_id))
                elif attribute['Key'] == 'قطع':
                    cursor.execute('UPDATE Book SET cover_size = %s WHERE id = %s', (attribute['Value'], product_id))
                elif attribute['Key'] == 'نوع جلد':
                    cursor.execute('UPDATE Book SET cover_material = %s WHERE id = %s', (attribute['Value'], product_id))
                elif attribute['Key'] == 'سایر توضیحات':
                    cursor.execute('UPDATE Book SET description = %s WHERE id = %s', (attribute['Value'], product_id))
    elif category == 'پازل':
        cursor.execute('INSERT INTO Puzzle (id, name_fa, name_en, brand_name_fa, url_code) VALUES (%s, %s, %s, %s, %s)', product_values)
        for attribute in attributes:
            if 'Value' in attribute and attribute['Value']:
                if attribute['Key'] == 'پازل':
                    cursor.execute('UPDATE Puzzle SET number_of_pieces = %s WHERE id = %s', (attribute['Value'], product_id))
                elif attribute['Key'] == 'نوع':
                    cursor.execute('UPDATE Puzzle SET type = %s WHERE id = %s', (attribute['Value'], product_id))
                elif attribute['Key'] == 'رده سنی':
                    if attribute['Value'] in ('کودک', 'خردسال'):
                        cursor.execute('UPDATE Puzzle SET age_rating = %s WHERE id = %s', ('کودک', product_id))
                    elif attribute['Value'] in ('نوجوان', 'بزرگسال'):
                        cursor.execute('UPDATE Puzzle SET age_rating = %s WHERE id = %s', ('بزرگسال', product_id))
                elif attribute['Key'] == 'خطر بلعیدن':
                    cursor.execute('UPDATE Puzzle SET danger_of_swallow = %s WHERE id = %s', (attribute['Value'], product_id))
                elif attribute['Key'] == 'وزن':
                    cursor.execute('UPDATE Puzzle SET weight = %s WHERE id = %s', (attribute['Value'], product_id))
                elif attribute['Key'] == 'وزن بسته‌بندی':
                    cursor.execute('UPDATE Puzzle SET package_weight = %s WHERE id = %s', (attribute['Value'], product_id))
                elif attribute['Key'] == 'ابعاد':
                    cursor.execute('UPDATE Puzzle SET dimensions = %s WHERE id = %s', (attribute['Value'], product_id))
                elif attribute['Key'] == 'ابعاد بسته‌بندی':
                    cursor.execute('UPDATE Puzzle SET package_dimensions = %s WHERE id = %s', (attribute['Value'], product_id))
                elif attribute['Key'] == 'سازنده':
                    cursor.execute('UPDATE Puzzle SET manufacturer = %s WHERE id = %s', (attribute['Value'], product_id))
                elif attribute['Key'] == 'محتویات بسته':
                    cursor.execute('UPDATE Puzzle SET items_in_package = %s WHERE id = %s', (attribute['Value'], product_id))
                elif attribute['Key'] == 'سایر توضیحات':
                    cursor.execute('UPDATE Puzzle SET description = %s WHERE id = %s', (attribute['Value'], product_id))

    elif category == 'ماوس (موشواره)':
        cursor.execute('INSERT INTO Mouse (id, name_fa, name_en, brand_name_fa, url_code) VALUES (%s, %s, %s, %s, %s)', product_values)
        for attribute in attributes:
            if 'Value' in attribute and attribute['Value']:
                if attribute['Key'] == 'ابعاد':
                    cursor.execute('UPDATE Mouse SET dimensions = %s WHERE id = %s', (attribute['Value'], product_id))
                elif attribute['Key'] == 'وزن':
                    cursor.execute('UPDATE Mouse SET weight = %s WHERE id = %s', (attribute['Value'], product_id))
                elif attribute['Key'] == 'رنگ':
                    cursor.execute('UPDATE Mouse SET color = %s WHERE id = %s', (attribute['Value'], product_id))
                elif attribute['Key'] == 'تعداد کلیدها':
                    cursor.execute('UPDATE Mouse SET number_of_buttons = %s WHERE id = %s', (attribute['Value'], product_id))
                elif attribute['Key'] == 'کلید روشن و خاموش':
                    cursor.execute('UPDATE Mouse SET power_button = %s WHERE id = %s', (attribute['Value'], product_id))
                elif attribute['Key'] == 'قابلیت کارکردن با هر دو دست':
                    cursor.execute('UPDATE Mouse SET two_handed = %s WHERE id = %s', (attribute['Value'], product_id))
                elif attribute['Key'] == 'نوع اتصال':
                    cursor.execute('UPDATE Mouse SET connection_type = %s WHERE id = %s', (attribute['Value'], product_id))
                elif attribute['Key'] == 'نوع رابط':
                    cursor.execute('UPDATE Mouse SET connection_port = %s WHERE id = %s', (attribute['Value'], product_id))
                elif attribute['Key'] == 'جنس کابل':
                    cursor.execute('UPDATE Mouse SET cable_material = %s WHERE id = %s', (attribute['Value'], product_id))
                elif attribute['Key'] == 'طول کابل':
                    cursor.execute('UPDATE Mouse SET cable_length = %s WHERE id = %s', (attribute['Value'], product_id))
                elif attribute['Key'] == 'نوع حسگر':
                    cursor.execute('UPDATE Mouse SET sensor_type = %s WHERE id = %s', (attribute['Value'], product_id))
                elif attribute['Key'] == 'دقت':
                    cursor.execute('UPDATE Mouse SET accuracy = %s WHERE id = %s', (attribute['Value'], product_id))
                elif attribute['Key'] == 'محدوده دقت':
                    cursor.execute('UPDATE Mouse SET accuracy_interval = %s WHERE id = %s', (attribute['Value'], product_id))
                elif attribute['Key'] == 'سازگار با سیستم‌عامل‌های':
                    cursor.execute('UPDATE Mouse SET compatible_os = %s WHERE id = %s', (attribute['Value'], product_id))
                elif attribute['Key'] == 'سایر قابلیت‌ها':
                    cursor.execute('UPDATE Mouse SET other_capabilities = %s WHERE id = %s', (attribute['Value'], product_id))
    elif category == 'کیبورد (صفحه کلید)':
        cursor.execute('INSERT INTO Keyboard (id, name_fa, name_en, brand_name_fa, url_code) VALUES (%s, %s, %s, %s, %s)', product_values)
        for attribute in attributes:
            if 'Value' in attribute and attribute['Value']:
                if attribute['Key'] == 'ابعاد':
                    cursor.execute('UPDATE Keyboard SET dimensions = %s WHERE id = %s', (attribute['Value'], product_id))
                elif attribute['Key'] == 'وزن':
                    cursor.execute('UPDATE Keyboard SET weight = %s WHERE id = %s', (attribute['Value'], product_id))
                elif attribute['Key'] == 'تعداد کلیدها':
                    cursor.execute('UPDATE Keyboard SET number_of_buttons = %s WHERE id = %s', (attribute['Value'], product_id))
                elif attribute['Key'] == 'عمر یا ضربه‌پذیری کلیدها':
                    cursor.execute('UPDATE Keyboard SET keys_lifetime = %s WHERE id = %s', (attribute['Value'], product_id))
                elif attribute['Key'] == 'کلید روشن و خاموش':
                    cursor.execute('UPDATE Keyboard SET power_button = %s WHERE id = %s', (attribute['Value'], product_id))
                elif attribute['Key'] == 'مقاوم در برابر گرد و غبار':
                    cursor.execute('UPDATE Keyboard SET dust_proof = %s WHERE id = %s', (attribute['Value'], product_id))
                elif attribute['Key'] == 'حروف حک شده فارسی':
                    cursor.execute('UPDATE Keyboard SET persian_letters = %s WHERE id = %s', (attribute['Value'], product_id))
                elif attribute['Key'] == 'چراغ‌ پس زمینه صفحه کلید':
                    cursor.execute('UPDATE Keyboard SET backlight = %s WHERE id = %s', (attribute['Value'], product_id))
                elif attribute['Key'] == 'همراه با ماوس':
                    cursor.execute('UPDATE Keyboard SET has_mouse = %s WHERE id = %s', (attribute['Value'], product_id))
                elif attribute['Key'] == 'تاچ پد':
                    cursor.execute('UPDATE Keyboard SET has_touchpad = %s WHERE id = %s', (attribute['Value'], product_id))
                elif attribute['Key'] == 'قابل استفاده با هر دو دست':
                    cursor.execute('UPDATE Keyboard SET two_handed = %s WHERE id = %s', (attribute['Value'], product_id))
                elif attribute['Key'] == 'نوع اتصال':
                    cursor.execute('UPDATE Keyboard SET connection_type = %s WHERE id = %s', (attribute['Value'], product_id))
                elif attribute['Key'] == 'نوع رابط':
                    cursor.execute('UPDATE Keyboard SET connection_port = %s WHERE id = %s', (attribute['Value'], product_id))
                elif attribute['Key'] == 'منبع تغذیه':
                    cursor.execute('UPDATE Keyboard SET power_source = %s WHERE id = %s', (attribute['Value'], product_id))
                elif attribute['Key'] == 'طول کابل':
                    cursor.execute('UPDATE Keyboard SET cable_length = %s WHERE id = %s', (attribute['Value'], product_id))
                elif attribute['Key'] == 'دقت':
                    cursor.execute('UPDATE Keyboard SET accuracy = %s WHERE id = %s', (attribute['Value'], product_id))
                elif attribute['Key'] == 'محدوده دقت':
                    cursor.execute('UPDATE Keyboard SET accuracy_interval = %s WHERE id = %s', (attribute['Value'], product_id))
    elif category == 'محافظ صفحه نمایش گوشی':
        cursor.execute('INSERT INTO ScreenProtector (id, name_fa, name_en, brand_name_fa, url_code) VALUES (%s, %s, %s, %s, %s)', product_values)
        for attribute in attributes:
            if 'Value' in attribute and attribute['Value']:
                if attribute['Key'] == 'مناسب برای گوشی های':
                    cursor.execute('UPDATE ScreenProtector SET device_model = %s WHERE id = %s', (attribute['Value'], product_id))
                elif attribute['Key'] == 'نوع':
                    cursor.execute('UPDATE ScreenProtector SET type = %s WHERE id = %s', (attribute['Value'], product_id))
                elif attribute['Key'] == 'ضخامت':
                    cursor.execute('UPDATE ScreenProtector SET thickness = %s WHERE id = %s', (attribute['Value'], product_id))
                elif attribute['Key'] == 'مقاوم در برابر ضربه':
                    cursor.execute('UPDATE ScreenProtector SET impact_protection = %s WHERE id = %s', (attribute['Value'], product_id))
                elif attribute['Key'] == 'جلوگیری از ایجاد خط و خش':
                    cursor.execute('UPDATE ScreenProtector SET scratch_protection = %s WHERE id = %s', (attribute['Value'], product_id))
                elif attribute['Key'] == 'جلوگیری از انعکاس نور':
                    cursor.execute('UPDATE ScreenProtector SET anti_reflection = %s WHERE id = %s', (attribute['Value'], product_id))
                elif attribute['Key'] == 'قابلیت نصب آسان':
                    cursor.execute('UPDATE ScreenProtector SET easy_install = %s WHERE id = %s', (attribute['Value'], product_id))
                elif attribute['Key'] == 'دارای محافظ برای قسمت:':
                    cursor.execute('UPDATE ScreenProtector SET protection_part = %s WHERE id = %s', (attribute['Value'], product_id))
                elif attribute['Key'] == 'مشخصات دیگر':
                    cursor.execute('UPDATE ScreenProtector SET description = %s WHERE id = %s', (attribute['Value'], product_id))
    elif category == 'کیف و کاور گوشی':
        cursor.execute('INSERT INTO Cover (id, name_fa, name_en, brand_name_fa, url_code) VALUES (%s, %s, %s, %s, %s)', product_values)
        for attribute in attributes:
            if 'Value' in attribute and attribute['Value']:
                if attribute['Key'] == 'مناسب برای گوشی موبایل':
                    cursor.execute('UPDATE Cover SET device_model = %s WHERE id = %s', (attribute['Value'], product_id))
                elif attribute['Key'] == 'نوع':
                    cursor.execute('UPDATE Cover SET type = %s WHERE id = %s', (attribute['Value'], product_id))
                elif attribute['Key'] == 'ساختار':
                    cursor.execute('UPDATE Cover SET structure = %s WHERE id = %s', (attribute['Value'], product_id))
                elif attribute['Key'] == 'جنس':
                    cursor.execute('UPDATE Cover SET material = %s WHERE id = %s', (attribute['Value'], product_id))
                elif attribute['Key'] == 'وزن':
                    cursor.execute('UPDATE Cover SET weight = %s WHERE id = %s', (attribute['Value'], product_id))
                elif attribute['Key'] == 'ابعاد':
                    cursor.execute('UPDATE Cover SET dimensions = %s WHERE id = %s', (attribute['Value'], product_id))
                elif attribute['Key'] == 'سطح پوشش':
                    if attribute['Value'] == 'لبه بالایی':
                        cursor.execute('UPDATE Cover SET top_edge_protection = %s WHERE id = %s', ('بله', product_id))
                    elif attribute['Value'] == 'لبه پایینی':
                        cursor.execute('UPDATE Cover SET bottom_edge_protection = %s WHERE id = %s', ('بله', product_id))
                    elif attribute['Value'] == 'لبه راست':
                        cursor.execute('UPDATE Cover SET right_edge_protection = %s WHERE id = %s', ('بله', product_id))
                    elif attribute['Value'] == 'لبه چپ':
                        cursor.execute('UPDATE Cover SET left_edge_protection = %s WHERE id = %s', ('بله', product_id))
                    elif attribute['Value'] == 'قاب پشتی':
                        cursor.execute('UPDATE Cover SET back_panel_protection = %s WHERE id = %s', ('بله', product_id))
                    elif attribute['Value'] == 'حفاظت از دکمه‌ها':
                        cursor.execute('UPDATE Cover SET buttons_protection = %s WHERE id = %s', ('بله', product_id))
                elif attribute['Key'] == 'قابلیت‌های ویژه':
                    cursor.execute('UPDATE Cover SET special_capabilities = %s WHERE id = %s', (attribute['Value'], product_id))
                elif attribute['Key'] == 'سایر توضیحات':
                    cursor.execute('UPDATE Cover SET description = %s WHERE id = %s', (attribute['Value'], product_id))


# Loading CSV file
data = pd.read_csv(buying_history_path)

# Storing CSV file into the database
for i in range(data.shape[0]):
    order_id = int(data['ID_Order'][i])
    customer_id = int(data['ID_Customer'][i])
    item_id = int(data['ID_Item'][i])
    cart_finalize_datetime = data['DateTime_CartFinalize'][i]
    order_gross_amount = int(data['Amount_Gross_Order'][i])
    city = data['city_name_fa'][i]
    item_quantity = int(data['Quantity_item'][i])
    values = (order_id, customer_id, item_id, cart_finalize_datetime, order_gross_amount, city, item_quantity)
    cursor.execute('INSERT INTO Buying_History (order_id, customer_id, item_id, cart_finalize_datetime, order_gross_amount, city, item_quantity) VALUES (%s, %s, %s, %s, %s, %s, %s)', values)

# Putting an index on Order Finalize Datetime
index_name = 'date_index'
cursor.execute('CREATE INDEX ' + index_name + ' ON Buying_History (cart_finalize_datetime)')


# Committing all the changes into the database
digikala.commit()


# Defining a function to analyse whether our sample queries use the index
def is_using_index(cursor_result):
    if cursor_result[0][6] == index_name:
        print('Our index was USED!')
        return True

    print('Our index was NOT USED!')
    return False


# Executing sample queries
cursor.execute('EXPLAIN SELECT cart_finalize_datetime, COUNT(*) FROM Buying_History GROUP BY cart_finalize_datetime HAVING COUNT(*) > 1')
is_using_index(cursor.fetchall())

cursor.execute('EXPLAIN SELECT AVG(order_gross_amount) FROM Buying_History WHERE item_quantity > 1 GROUP BY cart_finalize_datetime HAVING COUNT(*) > 2')
is_using_index(cursor.fetchall())

cursor.execute('EXPLAIN SELECT DISTINCT cart_finalize_datetime FROM Buying_History')
is_using_index(cursor.fetchall())

cursor.execute('EXPLAIN SELECT cart_finalize_datetime FROM Buying_History ORDER BY cart_finalize_datetime ASC')
is_using_index(cursor.fetchall())

cursor.execute('EXPLAIN SELECT id, cart_finalize_datetime FROM Buying_History WHERE cart_finalize_datetime >= "2018-05-05"')
is_using_index(cursor.fetchall())